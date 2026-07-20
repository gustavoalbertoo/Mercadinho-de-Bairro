"""
Gerador de dados para as tabelas Vendas e Estoque da planilha do Mercadinho do Bairro
--------------------------------------------------------------------------------------------------------------------------------------------------------
Este script abre a planilha já criada (mercadinho.xlsx), lê a lista de produtos que está na aba "Produtos" e gera dados aleatórios para duas abas novas:
Vendas: um histórico de vendas ao longo do ano, uma linha por venda.
Estoque: as movimentações de estoque. Toda venda gera uma SAÍDA e além disso o script cria ENTRADAS aleatórias para reposição de estoque.

A quantidade de vendas e de entradas geradas é controlada pelas constantes logo abaixo, então é só mudar os números para gerar mais ou menos dados.

IMPORTANTE: Precisa instalar as bibliotecas utilizadas.

Qualquer dúvida, contato: gustavo2020alberto@gmail.com
"""

import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#CONFIGURAÇÕES: Quantidade de dados que serão gerados
QUANTIDADE_DE_VENDAS = 11000         #quantas vendas aleatórias serão criadas (~30/dia)
DATA_INICIAL = date(2025, 8, 1)      #início do período simulado
DATA_FINAL = date(2026, 7, 19)       #fim do período simulado 

ARQUIVO_PLANILHA = "..." #caminho/localização do arquivo .xlsx

#Deixa os resultados sempre iguais entre execuções. Para gerar um conjunto de dados diferente a cada vez, é só remover esta linha abaixo
random.seed(42)

#Estilos usados nas duas tabelas
FONTE_PADRAO = "Calibri"
COR_CABECALHO = "2E7D32"
fonte_titulo = Font(name=FONTE_PADRAO, size=16, bold=True, color=COR_CABECALHO)
fonte_cabecalho = Font(name=FONTE_PADRAO, size=11, bold=True, color="FFFFFF")
fonte_normal = Font(name=FONTE_PADRAO, size=11)
preenchimento_cabecalho = PatternFill("solid", fgColor=COR_CABECALHO)
preenchimento_zebra = PatternFill("solid", fgColor="F1F8E9")
borda_fina = Border(*(Side(style="thin", color="CCCCCC") for _ in range(4)))


def formatar_cabecalho(planilha, linha, total_colunas):
    for coluna in range(1, total_colunas + 1):
        celula = planilha.cell(row=linha, column=coluna)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda_fina


def listrar_linhas(planilha, primeira_linha, ultima_linha, total_colunas):
    for linha in range(primeira_linha, ultima_linha + 1):
        cor = preenchimento_zebra if linha % 2 == 0 else PatternFill(fill_type=None)
        for coluna in range(1, total_colunas + 1):
            celula = planilha.cell(row=linha, column=coluna)
            celula.font = fonte_normal
            celula.fill = cor
            celula.border = borda_fina


def sortear_data_aleatoria(data_inicio, data_fim):
    #Sorteia um dia qualquer dentro do intervalo data_inicio e data_fim
    intervalo_em_dias = (data_fim - data_inicio).days
    dias_sorteados = random.randint(0, intervalo_em_dias)
    return data_inicio + timedelta(days=dias_sorteados)


#1. Abrir a planilha e ler o catálogo de produtos
planilha = openpyxl.load_workbook(ARQUIVO_PLANILHA)
aba_produtos = planilha["Produtos"]
aba_funcionarios = planilha["Funcionarios"]

#Cada produto vira um dicionário simples
produtos = []
for linha in aba_produtos.iter_rows(min_row=4, max_row=25, values_only=False):
    codigo = linha[0].value
    if codigo is None:
        continue
    produtos.append({
        "codigo": codigo,
        "nome": linha[1].value,
        "categoria": linha[2].value,
        "preco_custo": linha[3].value,
        "preco_venda": linha[4].value,
    })

#Lista de códigos dos funcionários que atuam como operadores de caixa para sortear quem realizou cada venda
codigos_operadores_caixa = []
for linha in aba_funcionarios.iter_rows(min_row=4, max_row=10, values_only=False):
    codigo_funcionario = linha[0].value
    cargo_funcionario = linha[4].value
    if codigo_funcionario and cargo_funcionario and "Caixa" in cargo_funcionario:
        codigos_operadores_caixa.append(codigo_funcionario)

#2. Gerar as vendas aleatórias
#Cada venda sorteia: um dia do período, um produto do catálogo, uma quantidade pequena (a maioria das compras de mercadinho é de poucas unidades) e o operador de caixa que atendeu.
vendas_geradas = []
for numero_da_venda in range(1, QUANTIDADE_DE_VENDAS + 1):
    produto_vendido = random.choice(produtos)
    quantidade_vendida = random.randint(1, 8)
    data_da_venda = sortear_data_aleatoria(DATA_INICIAL, DATA_FINAL)
    operador = random.choice(codigos_operadores_caixa) if codigos_operadores_caixa else ""

    valor_total_venda = round(produto_vendido["preco_venda"] * quantidade_vendida, 2)
    custo_total_venda = round(produto_vendido["preco_custo"] * quantidade_vendida, 2)

    vendas_geradas.append({
        "numero_venda": f"V{numero_da_venda:05d}",
        "data": data_da_venda,
        "codigo_produto": produto_vendido["codigo"],
        "nome_produto": produto_vendido["nome"],
        "categoria": produto_vendido["categoria"],
        "quantidade": quantidade_vendida,
        "preco_unitario": produto_vendido["preco_venda"],
        "valor_total": valor_total_venda,
        "custo_total": custo_total_venda,
        "operador": operador,
    })

#Ordena as vendas por data, para o histórico ficar em ordem cronológica.
vendas_geradas.sort(key=lambda venda: venda["data"])

#3. Gerar as entradas de estoque (compras/reposição junto a fornecedores)
#Em vez de sortear entradas totalmente ao acaso, calculamos por produto quanto saiu nas vendas e geramos reposições suficientes para cobrir essa saída 
#Isso evita estoque negativo e imita o que uma loja de verdade faz: repor o que vende, em lotes de compra menores
TAMANHO_MINIMO_LOTE_COMPRA = 20     #menor quantidade comprada por vez de um fornecedor
TAMANHO_MAXIMO_LOTE_COMPRA = 70     #maior quantidade comprada por vez de um fornecedor
MARGEM_DE_SEGURANCA_ESTOQUE = 1.25  #repõe 25% a mais do que foi vendido, como colchão

fornecedores_possiveis = [
    "Distribuidora Boa Safra", "Atacadão Central", "Laticínios Vale Verde",
    "Hortifruti Popular", "Bebidas São Jorge", "Comercial Higiene Total",
]

#Soma quanto cada produto vendeu ao longo do ano, para saber quanto repor
quantidade_vendida_por_produto = {}
for venda in vendas_geradas:
    codigo = venda["codigo_produto"]
    quantidade_vendida_por_produto[codigo] = (
        quantidade_vendida_por_produto.get(codigo, 0) + venda["quantidade"]
    )

entradas_geradas = []
for produto in produtos:
    codigo_produto = produto["codigo"]
    estoque_inicial_do_produto = next(
        (linha[8].value for linha in aba_produtos.iter_rows(min_row=4, max_row=25)
         if linha[0].value == codigo_produto), 0
    )
    total_vendido_no_ano = quantidade_vendida_por_produto.get(codigo_produto, 0)

    #Quanto precisa entrar no ano: o que foi vendido, menos o que já tinha de estoque inicial, mais a margem de segurança
    quantidade_necessaria = round(total_vendido_no_ano * MARGEM_DE_SEGURANCA_ESTOQUE) - estoque_inicial_do_produto
    quantidade_necessaria = max(quantidade_necessaria, TAMANHO_MINIMO_LOTE_COMPRA)

    #Quebra a quantidade necessária em vários lotes de compra ao longo do ano, simulando reposições periódicas em vez de uma compra gigante só
    quantidade_restante = quantidade_necessaria
    while quantidade_restante > 0:
        lote = min(random.randint(TAMANHO_MINIMO_LOTE_COMPRA, TAMANHO_MAXIMO_LOTE_COMPRA), quantidade_restante)
        entradas_geradas.append({
            "data": sortear_data_aleatoria(DATA_INICIAL, DATA_FINAL),
            "codigo_produto": produto["codigo"],
            "nome_produto": produto["nome"],
            "quantidade": lote,
            "fornecedor": random.choice(fornecedores_possiveis),
        })
        quantidade_restante -= lote

#4. Escrever a aba "Vendas"
aba_vendas = planilha.create_sheet("Vendas")
aba_vendas["A1"] = "Histórico de Vendas — Mercadinho do Bairro"
aba_vendas["A1"].font = fonte_titulo
aba_vendas.merge_cells("A1:J1")

cabecalho_vendas = [
    "Nº Venda", "Data", "Código Produto", "Produto", "Quantidade",
    "Preço Unitário", "Valor Total", "Custo Total", "Operador de Caixa", "Categoria",
]
linha_cabecalho_vendas = 3
for indice, titulo in enumerate(cabecalho_vendas, start=1):
    aba_vendas.cell(row=linha_cabecalho_vendas, column=indice, value=titulo)
formatar_cabecalho(aba_vendas, linha_cabecalho_vendas, len(cabecalho_vendas))

linha = linha_cabecalho_vendas + 1
primeira_linha_vendas = linha
for venda in vendas_geradas:
    aba_vendas.cell(row=linha, column=1, value=venda["numero_venda"])
    aba_vendas.cell(row=linha, column=2, value=venda["data"]).number_format = "DD/MM/YYYY"
    aba_vendas.cell(row=linha, column=3, value=venda["codigo_produto"])
    aba_vendas.cell(row=linha, column=4, value=venda["nome_produto"])
    aba_vendas.cell(row=linha, column=5, value=venda["quantidade"])
    aba_vendas.cell(row=linha, column=6, value=venda["preco_unitario"]).number_format = "R$ #,##0.00"
    aba_vendas.cell(row=linha, column=7, value=venda["valor_total"]).number_format = "R$ #,##0.00"
    aba_vendas.cell(row=linha, column=8, value=venda["custo_total"]).number_format = "R$ #,##0.00"
    aba_vendas.cell(row=linha, column=9, value=venda["operador"])
    aba_vendas.cell(row=linha, column=10, value=venda["categoria"])
    linha += 1
ultima_linha_vendas = linha - 1

listrar_linhas(aba_vendas, primeira_linha_vendas, ultima_linha_vendas, len(cabecalho_vendas))
for letra, largura in {"A": 10, "B": 12, "C": 14, "D": 22, "E": 11,
                        "F": 14, "G": 13, "H": 13, "I": 17, "J": 14}.items():
    aba_vendas.column_dimensions[letra].width = largura
aba_vendas.freeze_panes = "A4"

#5. Escrever a aba "Estoque" (saídas vindas das vendas + entradas de compra)
aba_estoque = planilha.create_sheet("Estoque")
aba_estoque["A1"] = "Movimentação de Estoque — Mercadinho do Bairro"
aba_estoque["A1"].font = fonte_titulo
aba_estoque.merge_cells("A1:G1")

cabecalho_estoque = [
    "Data", "Tipo", "Código Produto", "Produto", "Quantidade", "Origem/Destino", "Documento",
]
linha_cabecalho_estoque = 3
for indice, titulo in enumerate(cabecalho_estoque, start=1):
    aba_estoque.cell(row=linha_cabecalho_estoque, column=indice, value=titulo)
formatar_cabecalho(aba_estoque, linha_cabecalho_estoque, len(cabecalho_estoque))

#Junta entradas e saídas em uma única lista de movimentações e ordena por data, assim a aba de Estoque fica como um "extrato" cronológico
movimentacoes = []
for entrada in entradas_geradas:
    movimentacoes.append({
        "data": entrada["data"],
        "tipo": "Entrada",
        "codigo_produto": entrada["codigo_produto"],
        "nome_produto": entrada["nome_produto"],
        "quantidade": entrada["quantidade"],
        "origem_destino": entrada["fornecedor"],
        "documento": "Nota de Compra",
    })
for venda in vendas_geradas:
    movimentacoes.append({
        "data": venda["data"],
        "tipo": "Saída",
        "codigo_produto": venda["codigo_produto"],
        "nome_produto": venda["nome_produto"],
        "quantidade": venda["quantidade"],
        "origem_destino": "Venda ao cliente",
        "documento": venda["numero_venda"],
    })
movimentacoes.sort(key=lambda movimento: movimento["data"])

linha = linha_cabecalho_estoque + 1
primeira_linha_estoque = linha
for movimento in movimentacoes:
    aba_estoque.cell(row=linha, column=1, value=movimento["data"]).number_format = "DD/MM/YYYY"
    aba_estoque.cell(row=linha, column=2, value=movimento["tipo"])
    aba_estoque.cell(row=linha, column=3, value=movimento["codigo_produto"])
    aba_estoque.cell(row=linha, column=4, value=movimento["nome_produto"])
    quantidade_com_sinal = movimento["quantidade"]  # a coluna guarda sempre um número positivo
    aba_estoque.cell(row=linha, column=5, value=quantidade_com_sinal)
    aba_estoque.cell(row=linha, column=6, value=movimento["origem_destino"])
    aba_estoque.cell(row=linha, column=7, value=movimento["documento"])
    linha += 1
ultima_linha_estoque = linha - 1

listrar_linhas(aba_estoque, primeira_linha_estoque, ultima_linha_estoque, len(cabecalho_estoque))
for letra, largura in {"A": 12, "B": 10, "C": 14, "D": 22, "E": 11,
                        "F": 20, "G": 14}.items():
    aba_estoque.column_dimensions[letra].width = largura
aba_estoque.freeze_panes = "A4"

#Destaca "Entrada" em verde e "Saída" em vermelho para leitura rápida
verde_entrada = Font(name=FONTE_PADRAO, size=11, color="2E7D32", bold=True)
vermelho_saida = Font(name=FONTE_PADRAO, size=11, color="C62828", bold=True)
for linha in range(primeira_linha_estoque, ultima_linha_estoque + 1):
    celula_tipo = aba_estoque.cell(row=linha, column=2)
    celula_tipo.font = verde_entrada if celula_tipo.value == "Entrada" else vermelho_saida

#6. Salvar
planilha.save(ARQUIVO_PLANILHA)

